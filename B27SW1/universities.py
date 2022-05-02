import requests
from bs4 import BeautifulSoup
import bs4
import openpyxl
import wordpress
def get_html_text(url):
    try:
        r = requests.get(url, timeout=40)
        r.raise_for_status()
        r.encoding = r.apparent_encoding
        return r.text     
    except:
        return ""
 
def fill_univ_list(ulist, html):
    soup = BeautifulSoup(html, "html.parser")     
    for tr in soup.find('tbody').children:         
        if isinstance(tr, bs4.element.Tag):             
            tds = tr('td')             
            ulist.append([tds[0].text.strip(),tds[1].text.strip(),tds[2].text.strip(),tds[4].text.strip(),tds[5].text.strip()])
 
def print_univ_list(ulist, num):
    wb = openpyxl.Workbook()
    ws = wb.create_sheet('sheet1')
    
    print("{:^10}\t{:^6}\t{:^10}\t{:^10}\t{:^10}".format("排名","学校","省市","得分","教学层次",chr(12288)))
    for i in range(num):         
        u = ulist[i]
        ws.cell(i+1,1).value = u[0]
        ws.cell(i+1,2).value = u[1]
        ws.cell(i+1,3).value = u[2]
        ws.cell(i+1,4).value = u[3]
        ws.cell(i+1,5).value = u[4]
        print(u[3])  
        print("{:^10}\t{:^10}\t{:^10}\t{:^12}\t{:^12}".format(u[0],u[1],u[2],u[3],u[4],chr(12288)))
        wb.save('univer.xlsx')
      
        wordpress.postsimple(u[2], u[3], 'public')

 
def main():
    uinfo = []
    url = 'https://www.shanghairanking.cn/rankings/bcur/2021'    
    html = get_html_text(url)
    fill_univ_list(uinfo,html)
    print_univ_list(uinfo,30)
    
    
main()